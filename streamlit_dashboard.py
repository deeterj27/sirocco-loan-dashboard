import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from dateutil.relativedelta import relativedelta
import numpy as np
import os

st.set_page_config(page_title="Sirocco I LP Portfolio Dashboard", layout="wide", initial_sidebar_state="expanded")

# Custom CSS for Sirocco branding
st.markdown("""
<style>
    /* Main background */
    .stApp {
        background-color: #1a1a1a;
    }
    
    /* Headers */
    h1, h2, h3 {
        color: #FFFFFF !important;
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'Roboto', 'Helvetica', 'Arial', sans-serif;
    }
    
    /* Metrics */
    [data-testid="metric-container"] {
        background-color: #2d2d2d;
        border: 1px solid #3d3d3d;
        padding: 1rem;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.3);
    }
    
    [data-testid="metric-container"] [data-testid="stMetricLabel"] {
        color: #FDB813 !important;
        font-weight: 600;
        font-size: 0.9rem;
    }
    
    [data-testid="metric-container"] [data-testid="stMetricValue"] {
        color: #FFFFFF !important;
        font-weight: 700;
    }
    
    /* Sidebar */
    .css-1d391kg, [data-testid="stSidebar"] {
        background-color: #242424;
    }
    
    .css-1d391kg .stMarkdown, [data-testid="stSidebar"] .stMarkdown {
        color: #FFFFFF;
    }
    
    /* Info boxes */
    .stAlert {
        background-color: #2d2d2d;
        color: #FFFFFF;
        border: 1px solid #3d3d3d;
    }
    
    /* Tables */
    .dataframe {
        background-color: #2d2d2d !important;
        color: #FFFFFF !important;
    }
    
    .dataframe th {
        background-color: #FDB813 !important;
        color: #1a1a1a !important;
        font-weight: 600;
        border: none !important;
    }
    
    .dataframe td {
        background-color: #2d2d2d !important;
        color: #FFFFFF !important;
        border-color: #3d3d3d !important;
    }
    
    .dataframe tr:hover td {
        background-color: #3d3d3d !important;
    }
    
    /* Expander */
    .streamlit-expanderHeader {
        background-color: #2d2d2d;
        color: #FDB813 !important;
        border-radius: 8px;
    }
    
    .streamlit-expanderContent {
        background-color: #242424;
        border: 1px solid #3d3d3d;
    }
    
    /* File uploader */
    [data-testid="stFileUploadDropzone"] {
        background-color: #2d2d2d;
        border: 2px dashed #FDB813;
    }
    
    /* Checkbox */
    .stCheckbox label {
        color: #FFFFFF !important;
    }
    
    /* General text */
    .stMarkdown, .stText, p, span, div {
        color: #FFFFFF;
    }
    
    /* Column headers with Sirocco yellow */
    .css-1kyxreq {
        color: #FDB813 !important;
    }
    
    /* Buttons */
    .stButton button {
        background-color: #FDB813;
        color: #1a1a1a;
        font-weight: 600;
        border: none;
        border-radius: 4px;
        transition: all 0.3s;
    }
    
    .stButton button:hover {
        background-color: #fcc944;
        transform: translateY(-1px);
        box-shadow: 0 4px 8px rgba(253, 184, 19, 0.3);
    }
</style>
""", unsafe_allow_html=True)

# Helper functions
def safe_float(value):
    """Safely convert a value to float"""
    try:
        if isinstance(value, str):
            if value.lower() in ['interest only', 'n/a', '']:
                return 0.0
            value = value.replace('$', '').replace(',', '')
        return float(value)
    except (ValueError, TypeError):
        return 0.0

def excel_date_to_datetime(serial_date):
    """Convert Excel serial date to datetime"""
    if pd.isna(serial_date):
        return pd.NaT
    if isinstance(serial_date, datetime):
        return serial_date
    if isinstance(serial_date, str):
        try:
            return pd.to_datetime(serial_date)
        except:
            try:
                serial_date = float(serial_date)
            except:
                return pd.NaT
    try:
        return pd.to_datetime('1899-12-30') + pd.to_timedelta(serial_date, unit='D')
    except:
        return pd.NaT

def format_currency(value):
    """Format value as currency"""
    return f"${value:,.2f}" if pd.notna(value) and value != 0 else "$0.00"

def format_percent(value):
    """Format value as percentage"""
    return f"{value:.2%}" if pd.notna(value) and value != 0 else "0.00%"

def get_cell_value(sheet, locations, default=None):
    """Try to get value from multiple cell locations"""
    for location in locations:
        if sheet[location].value is not None:
            return sheet[location].value
    return default

def process_life_settlement_data(ls_file):
    """Process Life Settlement Excel file and return summary data"""
    try:
        ls_wb = load_workbook(ls_file, data_only=True)
        
        if 'Valuation Summary' not in ls_wb.sheetnames or 'Premium Stream' not in ls_wb.sheetnames:
            return None
        
        val_sheet = ls_wb['Valuation Summary']
        premium_sheet = ls_wb['Premium Stream']
        
        policies = []
        
        for row in range(3, 200):
            policy_id_cell = val_sheet[f'B{row}']
            if not policy_id_cell.value:
                break
                
            try:
                # Get NDB value first
                ndb_cell_value = val_sheet[f'V{row}'].value
                ndb_value = safe_float(str(ndb_cell_value or '0').replace('$', '').replace(',', ''))
                
                # If NDB is 0, check for Face Amount column (try common locations)
                if ndb_value == 0:
                    # Try column W first (next to V)
                    face_cell_value = val_sheet[f'W{row}'].value
                    face_amount = safe_float(str(face_cell_value or '0').replace('$', '').replace(',', ''))
                    if face_amount == 0:
                        # Try other possible columns for Face Amount
                        for col in ['X', 'Y', 'U', 'T']:
                            face_cell_value = val_sheet[f'{col}{row}'].value
                            face_amount = safe_float(str(face_cell_value or '0').replace('$', '').replace(',', ''))
                            if face_amount > 0:
                                break
                    if face_amount > 0:
                        ndb_value = face_amount
                
                policy_data = {
                    'Policy_ID': str(policy_id_cell.value),
                    'Insured_ID': str(val_sheet[f'C{row}'].value or ''),
                    'Name': str(val_sheet[f'D{row}'].value or ''),
                    'Age': safe_float(val_sheet[f'F{row}'].value),
                    'Gender': str(val_sheet[f'G{row}'].value or ''),
                    'NDB': ndb_value,
                    'Valuation': safe_float(str(val_sheet[f'Z{row}'].value or '0').replace('$', '').replace(',', '')),
                    'Cost_Basis': safe_float(str(val_sheet[f'AB{row}'].value or '0').replace('$', '').replace(',', '')),
                    'Remaining_LE': safe_float(val_sheet[f'AC{row}'].value),
                }
                policies.append(policy_data)
            except:
                continue
        
        if len(policies) == 0:
            return None
        
        # Calculate summary statistics
        total_policies = len(policies)
        total_ndb = sum(p['NDB'] for p in policies)
        total_valuation = sum(p['Valuation'] for p in policies)
        total_cost_basis = sum(p['Cost_Basis'] for p in policies)
        
        valid_ages = [p['Age'] for p in policies if p['Age'] > 0]
        avg_age = sum(valid_ages) / len(valid_ages) if valid_ages else 0
        
        male_count = sum(1 for p in policies if 'male' in p['Gender'].lower() and 'female' not in p['Gender'].lower())
        female_count = sum(1 for p in policies if 'female' in p['Gender'].lower())
        male_percentage = (male_count / (male_count + female_count)) * 100 if (male_count + female_count) > 0 else 0
        
        valid_les = [p['Remaining_LE'] for p in policies if p['Remaining_LE'] > 0]
        avg_remaining_le = sum(valid_les) / len(valid_les) if valid_les else 0
        
        # Process monthly premiums
        monthly_premiums = {}
        policy_premiums = {}
        month_columns = ['M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X']
        
        month_headers = []
        for col in month_columns:
            header = premium_sheet[f'{col}2'].value
            if header:
                month_headers.append((col, str(header)))
        
        for col_letter, month_name in month_headers:
            month_total = 0
            for prem_row in range(3, len(policies) + 3):
                try:
                    lyric_id_cell = premium_sheet[f'B{prem_row}']
                    if lyric_id_cell.value:
                        lyric_id = str(lyric_id_cell.value)
                        premium_cell = premium_sheet[f'{col_letter}{prem_row}']
                        premium_val = safe_float(premium_cell.value) if premium_cell.value else 0
                        month_total += premium_val
                        
                        if lyric_id not in policy_premiums:
                            policy_premiums[lyric_id] = {}
                        policy_premiums[lyric_id][month_name] = premium_val
                except:
                    continue
            
            monthly_premiums[month_name] = month_total
        
        # Calculate policy-level metrics
        for policy in policies:
            policy_id = policy['Policy_ID']
            if policy_id in policy_premiums:
                annual_premium = sum(policy_premiums[policy_id].values())
                policy['Annual_Premium'] = annual_premium
                if policy['NDB'] > 0:
                    policy['Premium_Pct_Face'] = (annual_premium / policy['NDB']) * 100
                else:
                    policy['Premium_Pct_Face'] = 0
            else:
                policy['Annual_Premium'] = 0
                policy['Premium_Pct_Face'] = 0
        
        total_annual_premiums = sum(monthly_premiums.values())
        premiums_as_pct_face = (total_annual_premiums / total_ndb) * 100 if total_ndb > 0 else 0
        
        return {
            'policies': policies,
            'summary': {
                'total_policies': total_policies,
                'total_ndb': total_ndb,
                'total_valuation': total_valuation,
                'total_cost_basis': total_cost_basis,
                'avg_age': avg_age,
                'male_count': male_count,
                'female_count': female_count,
                'male_percentage': male_percentage,
                'avg_remaining_le': avg_remaining_le,
                'total_annual_premiums': total_annual_premiums,
                'premiums_as_pct_face': premiums_as_pct_face,
            },
            'monthly_premiums': monthly_premiums,
            'policy_premiums': policy_premiums
        }
        
    except:
        return None

# Main app

# Header with Sirocco branding
st.markdown("""
<div style='background-color: #1a1a1a; padding: 2rem 0; margin: -2rem -2rem 2rem -2rem; border-bottom: 4px solid #FDB813;'>
    <h1 style='text-align: center; color: #FFFFFF; font-size: 2.5rem; margin: 0;'>
        <span style='color: #FDB813;'>⚡</span> Sirocco I LP Portfolio Dashboard
    </h1>
    <p style='text-align: center; color: #999999; margin-top: 0.5rem;'>Loan Participation & Life Settlement Portfolio Management System</p>
</div>
""", unsafe_allow_html=True)

# Add version indicator
st.info("Dashboard Version: 2.0 - WITH UNREALIZED GAIN/LOSS")

# File upload section
col1, col2, col3 = st.columns(3)
with col1:
    master_file = st.file_uploader("Upload Master Excel File", type=["xlsx"])
with col2:
    ls_file = st.file_uploader("Upload LS Portfolio File", type=["xlsx"])
with col3:
    remittance_file = st.file_uploader("Upload Monthly Remittance File (CSV or XLSX)", type=["csv", "xlsx"])

# Process LS data if uploaded
ls_data = None
if ls_file:
    ls_data = process_life_settlement_data(ls_file)
    if ls_data:
        st.success(f"✅ Life Settlement data loaded: {ls_data['summary']['total_policies']} policies, {format_currency(ls_data['summary']['total_ndb'])} face value")
    else:
        st.error("❌ Failed to load Life Settlement data. Please check file format.")

# Process loan data
if master_file:
    try:
        # Load workbook
        wb = load_workbook(master_file, data_only=True)
        
        # Get all loan sheets (sheets starting with '#')
        loan_sheets = [s for s in wb.sheetnames if s.startswith('#') and s != '#AddSheet']
        
        # Get as-of date from Dashboard
        dashboard_sheet = wb['Dashboard']
        as_of_date = dashboard_sheet['E3'].value
        if isinstance(as_of_date, str):
            as_of_date = pd.to_datetime(as_of_date)
        elif isinstance(as_of_date, (int, float)):
            as_of_date = excel_date_to_datetime(as_of_date)
        
        # Sidebar with Sirocco branding
        with st.sidebar:
            st.markdown("""
            <div style='text-align: center; padding: 1rem 0;'>
                <h2 style='color: #FDB813; margin: 0;'>⚡ Sirocco Partners</h2>
            </div>
            """, unsafe_allow_html=True)
            
            st.markdown(f"""
            <div style='background-color: #2d2d2d; padding: 1rem; border-radius: 8px; margin-bottom: 1rem;'>
                <p style='color: #FDB813; margin: 0; font-weight: 600;'>📅 Data as of</p>
                <p style='color: #FFFFFF; margin: 0; font-size: 1.2rem;'>{as_of_date.strftime('%B %d, %Y') if pd.notna(as_of_date) else 'Unknown'}</p>
            </div>
            """, unsafe_allow_html=True)
            
            st.markdown(f"""
            <div style='background-color: #2d2d2d; padding: 1rem; border-radius: 8px;'>
                <p style='color: #FDB813; margin: 0; font-weight: 600;'>📊 Total Loans</p>
                <p style='color: #FFFFFF; margin: 0; font-size: 1.2rem;'>{len(loan_sheets)}</p>
            </div>
            """, unsafe_allow_html=True)
        
        st.success(f"✅ Master file loaded: {len(loan_sheets)} loan sheets found")
        
        # Display LS data if available (but after all loan data)
        if ls_data:
            st.markdown("<h2 style='color: #FDB813; margin-top: 3rem; font-size: 2rem;'>🏥 Life Settlement Portfolio</h2>", unsafe_allow_html=True)
            
            # Add a debug message to confirm updates
            st.success("✅ UPDATED VERSION WITH UNREALIZED GAIN/LOSS")
            
            # Key Metrics Box with Unrealized Gain/Loss
            unrealized_gain_loss = ls_data['summary']['total_valuation'] - ls_data['summary']['total_cost_basis']
            gain_loss_pct = (unrealized_gain_loss / ls_data['summary']['total_cost_basis'] * 100) if ls_data['summary']['total_cost_basis'] > 0 else 0
            
            # Create 5 columns for the metrics
            col1, col2, col3, col4, col5 = st.columns(5)
            
            with col1:
                st.metric("Total Policies", ls_data['summary']['total_policies'])
            with col2:
                st.metric("Total NDB (Face Value)", format_currency(ls_data['summary']['total_ndb']))
            with col3:
                st.metric("Total Valuation", format_currency(ls_data['summary']['total_valuation']))
            with col4:
                st.metric("Cost Basis", format_currency(ls_data['summary']['total_cost_basis']))
            with col5:
                st.metric("Unrealized Gain/(Loss)", 
                          format_currency(unrealized_gain_loss),
                          delta=f"{gain_loss_pct:.1f}%",
                          delta_color="normal" if unrealized_gain_loss >= 0 else "inverse")
            
            # Additional metrics in a second row
            col6, col7, col8, col9 = st.columns(4)
            
            with col6:
                st.metric("Average Age", f"{ls_data['summary']['avg_age']:.1f} years")
            with col7:
                st.metric("% Male", f"{ls_data['summary']['male_percentage']:.1f}%")
            with col8:
                st.metric("Avg Remaining LE", f"{ls_data['summary']['avg_remaining_le']:.1f} months")
            with col9:
                st.metric("Premiums % of Face", f"{ls_data['summary']['premiums_as_pct_face']:.2f}%")
            
            # Monthly Premium Projections
            if ls_data['monthly_premiums']:
                st.markdown("<h3 style='color: #FDB813; margin-top: 2rem;'>💵 Monthly Premium Projections</h3>", unsafe_allow_html=True)
                
                # Create a grid of premium months
                premium_items = list(ls_data['monthly_premiums'].items())
                num_months = len(premium_items)
                cols_per_row = 6
                
                for i in range(0, num_months, cols_per_row):
                    month_batch = premium_items[i:i+cols_per_row]
                    cols = st.columns(len(month_batch))
                    
                    for j, (month, amount) in enumerate(month_batch):
                        with cols[j]:
                            st.metric(month, format_currency(amount))
            
            # Policy Details Table
            st.markdown("<h3 style='color: #FFFFFF; margin-top: 2rem; font-size: 1.4rem;'>📋 Policy Details</h3>", unsafe_allow_html=True)
            
            if ls_data['policies']:
                policies_df = pd.DataFrame(ls_data['policies'])
                
                # Calculate unrealized gain/loss for each policy
                policies_df['Unrealized_Gain_Loss'] = policies_df['Valuation'] - policies_df['Cost_Basis']
                policies_df['Gain_Loss_Pct'] = (policies_df['Unrealized_Gain_Loss'] / policies_df['Cost_Basis'] * 100).fillna(0)
                
                # Format for display
                display_policy_df = policies_df.copy()
                display_policy_df['NDB'] = display_policy_df['NDB'].apply(format_currency)
                display_policy_df['Cost_Basis'] = display_policy_df['Cost_Basis'].apply(format_currency)
                display_policy_df['Valuation'] = display_policy_df['Valuation'].apply(format_currency)
                display_policy_df['Unrealized_Gain_Loss'] = display_policy_df['Unrealized_Gain_Loss'].apply(format_currency)
                display_policy_df['Annual_Premium'] = display_policy_df['Annual_Premium'].apply(format_currency)
                display_policy_df['Premium_Pct_Face'] = display_policy_df['Premium_Pct_Face'].apply(lambda x: f"{x:.2f}%")
                
                # Rename columns
                display_policy_df = display_policy_df.rename(columns={
                    'Policy_ID': 'Policy ID',
                    'Name': 'Name',
                    'Age': 'Age',
                    'Gender': 'Gender',
                    'NDB': 'Face Value',
                    'Cost_Basis': 'Cost Basis',
                    'Valuation': 'Valuation',
                    'Unrealized_Gain_Loss': 'Unrealized Gain/(Loss)',
                    'Annual_Premium': 'Annual Premium',
                    'Premium_Pct_Face': 'Premium % Face'
                })
                
                # Select columns to display
                display_columns = ['Policy ID', 'Name', 'Age', 'Gender', 'Face Value', 'Valuation', 'Cost Basis', 'Unrealized Gain/(Loss)', 'Annual Premium', 'Premium % Face']
                st.dataframe(display_policy_df[display_columns], use_container_width=True, hide_index=True)
        
    except Exception as e:
        st.error(f"Error processing master file: {str(e)}")
        st.error("Please ensure the Excel file has the expected structure with loan sheets starting with '#'")
        
        with st.expander("Debug Information"):
            st.code(str(e))

else:
    # Landing page
    st.markdown("""
    <div style='text-align: center; padding: 3rem 0;'>
        <div style='font-size: 5rem; color: #FDB813;'>⚡</div>
        <h2 style='color: #FFFFFF; margin-top: 1rem;'>Welcome to the Sirocco I LP Portfolio Dashboard</h2>
        <p style='color: #999999; font-size: 1.2rem; margin-top: 1rem;'>
            Upload your Master Excel file to begin analyzing your portfolio
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # Show expected file structure
    with st.expander("📋 Expected Excel File Structure"):
        st.markdown("""
        <div style='color: #FFFFFF;'>
        The Master Excel file should contain:
        
        **Dashboard sheet**: Summary of all loans
        
        **Loan sheets**: Named with # prefix (e.g., #1, #2, etc.)
        
        Each loan sheet should have loan information in either:
        
        **Format 1** (Data in column B):
        - A2 or B2: Borrower name
        - B3: Original loan amount
        - B4: Annual interest rate
        - B5: Loan period in months
        - B6: Payment amount
        - B7: Loan start date
        
        **Format 2** (Data in column C):
        - A2 or B2: Borrower name
        - C3: Original loan amount (when B3 contains label)
        - C4: Annual interest rate
        - C5: Loan period in months
        - C6: Payment amount
        - C7: Loan start date
        
        **Amortization schedule** starting from row 11 with columns:
        - A: Month
        - B: Repayment number
        - C: Opening balance
        - D: Loan repayment
        - E: Interest charged
        - F: Capital repaid
        - G: Closing balance
        - J: Payment date
        - K: Amount paid
        </div>
        """, unsafe_allow_html=True)

# Footer
st.markdown("""
<div style='margin-top: 3rem; padding-top: 2rem; border-top: 1px solid #3d3d3d; text-align: center; color: #666666;'>
    <p>Sirocco Partners - Portfolio Management System</p>
</div>
""", unsafe_allow_html=True) 